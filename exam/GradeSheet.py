from openpyxl import Workbook
from openpyxl import load_workbook
from utils import *

class GradeSheet:
    """Process Teams Gradesheet"""

    def __init__(self, filename):
        self.filename = filename

    def read_students(self):
        workbook = load_workbook(filename = self.filename, data_only=True)
        worksheet = workbook['Students']

        header = { cell.value : idx for idx,cell in enumerate(worksheet[1]) }
        students = {}

        # read all lines
        for row in worksheet.iter_rows(min_row=2, min_col=1):
            values = {}
            for key, cell in zip(header.keys(), row):
                if key is not None:
                    values[key] = cell.value
            # skip None lines
            if values['First name'] is None:
                continue
            student =values['ID number']
            students[student] = values

        return students
    