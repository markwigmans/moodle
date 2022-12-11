'''
convert Moodle participation into spreadsheet
'''
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def get_data(filename, skip_header=True):
    '''get data from forum CSV file'''
    # utf-8-bom encoding
    with open(filename, 'r', encoding='utf_8_sig') as file:
        reader = csv.reader(file)
        if skip_header:
            next(reader)
        rows = []
        for row in reader:
            rows.append(row)
        return rows

def get_header(data):
    '''get header of CSV file'''
    header = {}
    for idx, item in enumerate(data[0]):
        header[item] = idx
    return data[1:],header

def get_students(data,header):
    '''retrieve dictonary with student as key and as value all his postings'''
    students = {}
    for row in data:
        student = normalize_key(row[header['userfullname']])
        students[student] = students.get(student, []) + [ merge(list(header.keys()), row) ]
    return students

def normalize_key(string):
    '''remove all whitespace characters'''
    return ''.join(string.split()).lower()

def merge(header,row):
    '''Merge header and row into dictonary'''
    return { header[i]: row[i].strip() for i in range(len(header)) }


def read_grade_sheet(filename):
    '''read participation part of grade spreadsheet'''
    workbook = load_workbook(filename = filename, data_only=True)
    worksheet = workbook['Quizes - Participation']

    header = [cell.value for cell in worksheet[1]]
    students = {}

    for row in worksheet.iter_rows(min_row=2, min_col=1):
        values = {}
        for key, cell in zip(header, row):
            if key is not None:
                values[key] = cell.value
        # skip None lines
        if values[header[0]] is None:
            continue
        student = normalize_key(f"{values[header[0]]} {values[header[1]]}")
        students[student] = values
    return students

def align_data(worksheet):
    '''Align all data of the given sheet'''
    for row in range(2,worksheet.max_row+1):
        for column in range(1,worksheet.max_column+1):
            worksheet.cell(row=row, column=column).alignment = Alignment(vertical='top', wrap_text=True)



def gen_sheet(filename, participation, sheets):
    '''Generate resulting Excel sheet'''
    workbook = Workbook()
    worksheet = workbook.active

    #
    # Create readme part
    #
    worksheet.title = 'Readme'
    worksheet.cell(row=1, column=1).value = "Paricipation overview"
    worksheet.cell(row=2, column=1).value = "Sheet 'overview' contains per student the number of posts"
    worksheet.cell(row=3, column=1).value = "Sheet 'posts' contains all posts of all students"

    row = 4
    for sheet,description,_ in sheets:
        row += 1
        worksheet.cell(row=row, column=1).value = f"{sheet} : {description}"

    #
    # Overview part
    #
    worksheet = workbook.create_sheet(title='Overview')

    #
    # fill first sheet with overall participation
    #
    colnames = ['First name', 'Surname', 'Marker'] + [ x for x,description,students in sheets] + ['Forums']
    for idx,header in enumerate(colnames):
        worksheet.cell(row=1, column=idx+1).value = header
        worksheet.cell(row=1, column=idx+1).alignment = Alignment(horizontal='center')

    # data part
    row = 1
    for student,value in overview.items():
        row += 1
        worksheet.cell(row=row, column=1).value = value['First name']
        worksheet.cell(row=row, column=2).value = value['Surname']
        worksheet.cell(row=row, column=3).value = value['Marker']
        column = 3
        for sheet,_,students in sheets:
            column += 1
            worksheet.cell(row=row, column=column).value = len(students[student]) if participation[student][sheet] else '-'
            worksheet.cell(row=row, column=column).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row, column=column+1).value = participation[student]['total']
        worksheet.auto_filter.ref = "C1:C" + str(worksheet.max_row)

    #
    # generate sheets with all posts
    #
    worksheet = workbook.create_sheet(title='Posts')
    worksheet.column_dimensions['E'].width = 40
    worksheet.column_dimensions['F'].width = 100

    headers = ['subject', 'message', 'wordcount']
    colnames = ['First name', 'Surname', 'Marker', 'Forum'] + headers
    row = 1
    for idx,header in enumerate(colnames):
        worksheet.cell(row=row, column=idx+1).value = header

    for student,value in overview.items():
        for sheet,_,students in sheets:
            # skip students without post, but also keep the same order as in overview
            if student not in students:
                continue
            for data in students[student]:
                row += 1
                worksheet.cell(row=row, column=1).value = value['First name']
                worksheet.cell(row=row, column=2).value = value['Surname']
                worksheet.cell(row=row, column=3).value = value['Marker']
                worksheet.cell(row=row, column=4).value = sheet
                column = 4
                for key in headers:
                    column += 1
                    worksheet.cell(row=row, column=column).value = data[key]
        worksheet.auto_filter.ref = "A1:D" + str(worksheet.max_row)
        align_data(worksheet)

    #
    # generate sheets with the data
    #
    for sheet,_,students in sheets:
        worksheet = workbook.create_sheet(title=sheet)
        worksheet.column_dimensions['D'].width = 40
        worksheet.column_dimensions['E'].width = 100

        # Header
        headers = ['subject', 'message', 'wordcount']
        colnames = ['First name', 'Surname', 'Marker'] + headers
        for idx,header in enumerate(colnames):
            worksheet.cell(row=1, column=idx+1).value = header

        # Data
        row = 1
        for student,data in overview.items():
            # skip students without post, but also keep the same order as in overview
            if student not in students:
                continue
            for value in students[student]:
                row += 1
                worksheet.cell(row=row, column=1).value = data['First name']
                worksheet.cell(row=row, column=2).value = data['Surname']
                worksheet.cell(row=row, column=3).value = data['Marker']
                column = 3
                for key in headers:
                    column += 1
                    worksheet.cell(row=row, column=column).value = value[key]
                    worksheet.cell(row=row, column=column).alignment = Alignment(wrap_text=True)
            worksheet.auto_filter.ref = "A1:C" + str(worksheet.max_row)
            align_data(worksheet)

    workbook.active = workbook['Overview']
    workbook.save(filename=filename)


def calc_participation(sheets):
    '''Calculate overall participation excel file'''
    participation = {}

    # create empty dictonaries for every student
    for student in overview:
        key = normalize_key(student)
        participation[key] = { 'total' : 0}

    for sheet,_,students in sheets:
        for student in overview:
            key = normalize_key(student)
            count = participation[key]['total']
            if key in students:
                count += 1
            participation[key].update({ sheet : key in students, 'total' : count })
    return participation

#
# Main application
#
overview = read_grade_sheet('bloc-516 - Fall 2022 - Exam Marks.xlsx')
data1,csv_header = get_header(get_data('discussion-1.csv', False))
students1 = get_students(data1,csv_header)
students2 = get_students(get_data('discussion-2.csv'),csv_header)

sheets_list = [
    ('P-1', 'Session 4 - Creating Money', students1),
    ("P-2", 'Session 6 - Securitization', students2)]
gen_sheet('participation.xlsx', calc_participation(sheets_list), sheets_list)
