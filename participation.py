import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment 

#
# get data from forum CSV file
#
def get_data(filename, skip_header=True):
    # utf-8-bom encoding
    with open(filename, 'r', encoding='utf_8_sig') as f:
        reader = csv.reader(f)
        if (skip_header): 
            next(reader)
        rows = []
        for row in reader:
            rows.append(row)
        return rows;

#
# get header of CSV file
#
def get_header(data):
    header = {}
    for idx, item in enumerate(data[0]):
        header[item] = idx
    return data[1:],header

#
#
#
def get_students(data):
    students = {}
    for row in data:
        student = normalize_key(row[header['userfullname']])
        students[student] = merge(list(header.keys()), row)
    return students;

#
# remove all whitespace characters
#
def normalize_key(string):
    return ''.join(string.split()).lower()

def merge(h,row):
    return { h[i]: row[i] for i in range(len(h)) }

#
# read participation part of grade spreadsheet
#
def read_grade_sheet(filename):
    wb = load_workbook(filename = filename, data_only=True)
    ws = wb['Quizes - Participation']

    header = [cell.value for cell in ws[1]]
    students = {}

    for row in ws.iter_rows(min_row=2, min_col=1):
        values = {}
        for key, cell in zip(header, row):
            if key is not None:
                values[key] = cell.value
        # skip None lines        
        if values[header[0]] is None:
            continue
        student = normalize_key(f"{values[header[0]]} {values[header[1]]}")
        students[student] = values;
    return students;

#
#
#
def gen_sheet(filename, participation, sheets):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Overview'

    #
    # fill first sheet with overall participation
    #
    colnames = ['First name', 'Surname', 'Marker'] + [ x for x,y in sheets] + ['Total']
    for idx,h in enumerate(colnames):
        ws.cell(row=1, column=idx+1).value = h
        ws.cell(row=1, column=idx+1).alignment = Alignment(horizontal='center')

    # data part
    for idx,student in enumerate(overview.keys()):
        row = idx+2
        ws.cell(row=row, column=1).value = overview[student]['First name']
        ws.cell(row=row, column=2).value = overview[student]['Surname']
        ws.cell(row=row, column=3).value = overview[student]['Marker']
        column = 4
        for sheet,students in sheets:
            ws.cell(row=row, column=column).value = 'X' if participation[student][sheet] else '-'
            ws.cell(row=row, column=column).alignment = Alignment(horizontal='center')
            column += 1
        ws.cell(row=row, column=column).value = participation[student]['total']

    #
    # generate sheets with the data
    #
    for sheet,students in sheets:
        ws = wb.create_sheet(title=sheet)
        # Header
        colnames = ['First name', 'Surname', 'Marker'] + [ x for x in header.keys()]
        for idx,h in enumerate(colnames):
            ws.cell(row=1, column=idx+1).value = h

        # Data
        for idx,student in enumerate(students):
            row = idx+2
            ws.cell(row=row, column=1).value = overview[student]['First name']
            ws.cell(row=row, column=2).value = overview[student]['Surname']
            ws.cell(row=row, column=3).value = overview[student]['Marker']
            column = 4
            for key in header.keys():
                ws.cell(row=row, column=column).value = students[student][key]
                column += 1

    wb.save(filename=filename)

#
#
# 
def calc_participation(sheets):
    participation = {}

    # create empty dictonaries for every student
    for student in overview.keys():
        key = normalize_key(student)
        participation[key] = { 'total' : 0}

    for sheet,students in sheets:
        for student in overview.keys():
            key = normalize_key(student)
            count = participation[key]['total']
            if key in students:
                count += 1
            participation[key].update({ sheet : key in students, 'total' : count })
    return participation        

#
# Main application
#
overview = read_grade_sheet('bloc-516 - Fall 2022 - Exam Marks.xlsx')
data1,header = get_header(get_data('discussion-1.csv', False))
data2 = get_data('discussion-2.csv')

students1 = get_students(data1)
students2 = get_students(data2)

sheets = [("P-1", students1),("P-2", students2)]
participation = calc_participation(sheets)
gen_sheet('participation.xlsx', participation, sheets);